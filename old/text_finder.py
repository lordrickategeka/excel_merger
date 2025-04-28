import os
import sys
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext
import threading
import queue

class FileSearchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("File String Searcher")
        self.root.geometry("800x600")
        self.root.minsize(800, 600)
        
        self.folder_path = tk.StringVar()
        self.search_string = tk.StringVar()
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        self.file_extensions = tk.StringVar(value=".txt,.xlsx,.xls,.csv")
        
        self.result_queue = queue.Queue()
        self.create_widgets()
        
    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Input Frame
        input_frame = ttk.LabelFrame(main_frame, text="Search Parameters", padding="10")
        input_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Folder selection
        folder_frame = ttk.Frame(input_frame)
        folder_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(folder_frame, text="Folder:").pack(side=tk.LEFT, padx=5)
        ttk.Entry(folder_frame, textvariable=self.folder_path, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(folder_frame, text="Browse...", command=self.browse_folder).pack(side=tk.LEFT, padx=5)
        
        # Search string
        search_frame = ttk.Frame(input_frame)
        search_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(search_frame, text="Search Text:").pack(side=tk.LEFT, padx=5)
        ttk.Entry(search_frame, textvariable=self.search_string, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # File extensions
        ext_frame = ttk.Frame(input_frame)
        ext_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(ext_frame, text="File Extensions:").pack(side=tk.LEFT, padx=5)
        ttk.Entry(ext_frame, textvariable=self.file_extensions, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Label(ext_frame, text="(comma-separated)").pack(side=tk.LEFT, padx=5)
        
        # Button frame
        button_frame = ttk.Frame(input_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(button_frame, text="Search", command=self.start_search).pack(side=tk.RIGHT, padx=5)
        
        # Results Area
        result_frame = ttk.LabelFrame(main_frame, text="Search Results", padding="10")
        result_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.result_text = scrolledtext.ScrolledText(result_frame, wrap=tk.WORD, width=80, height=20)
        self.result_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.result_text.config(state=tk.DISABLED)
        
        # Status bar
        status_bar = ttk.Frame(self.root)
        status_bar.pack(fill=tk.X, side=tk.BOTTOM, pady=2)
        
        ttk.Label(status_bar, textvariable=self.status_var).pack(side=tk.LEFT, padx=5)
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(status_bar, variable=self.progress_var, mode='indeterminate')
        self.progress_bar.pack(side=tk.RIGHT, padx=5, fill=tk.X, expand=True)
        
    def browse_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.folder_path.set(folder_selected)
    
    def start_search(self):
        folder = self.folder_path.get()
        search_text = self.search_string.get()
        extensions = self.file_extensions.get()
        
        if not folder or not search_text:
            self.status_var.set("Error: Please provide both folder and search text")
            return
        
        if not os.path.isdir(folder):
            self.status_var.set(f"Error: '{folder}' is not a valid directory")
            return
        
        # Clear previous results
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)
        self.result_text.config(state=tk.DISABLED)
        
        # Start progress bar
        self.progress_bar.start(10)
        self.status_var.set("Searching...")
        
        # Process file extensions
        file_exts = [ext.strip() if ext.strip().startswith('.') else f'.{ext.strip()}' 
                  for ext in extensions.split(',')]
        
        # Start search in a separate thread
        search_thread = threading.Thread(
            target=self.search_files_thread,
            args=(folder, search_text, file_exts),
            daemon=True
        )
        search_thread.start()
        
        # Start checking for results
        self.root.after(100, self.check_queue)
    
    def search_files_thread(self, folder_path, search_string, file_extensions):
        try:
            # Initial info message
            self.result_queue.put(("info", f"Searching for '{search_string}' in {folder_path}...\n"))
            self.result_queue.put(("info", f"File extensions: {', '.join(file_extensions)}\n"))
            self.result_queue.put(("info", "-" * 80 + "\n"))
            
            results = self.search_files_in_folder(folder_path, search_string, file_extensions)
            
            if not results:
                self.result_queue.put(("info", f"No matches found for '{search_string}'.\n"))
            else:
                self.result_queue.put(("info", f"Found matches in {len(results)} files:\n"))
                
                for file_path, matches in results.items():
                    rel_path = os.path.relpath(file_path, folder_path)
                    self.result_queue.put(("file", f"\n📄 {rel_path}\n"))
                    
                    for i, match in enumerate(matches, 1):
                        if isinstance(match, tuple) and len(match) == 2:
                            location, content = match
                            self.result_queue.put(("location", f"  {i}. Location: {location}\n"))
                            self.result_queue.put(("content", f"     Content: {content[:100]}{'...' if len(content) > 100 else ''}\n"))
                        else:
                            self.result_queue.put(("error", f"  {i}. {match}\n"))  # This is likely an error message
                    
                    self.result_queue.put(("info", "-" * 80 + "\n"))
                
                self.result_queue.put(("summary", f"Total files with matches: {len(results)}\n"))
        except Exception as e:
            self.result_queue.put(("error", f"Error occurred: {str(e)}\n"))
        finally:
            self.result_queue.put(("done", None))
    
    def check_queue(self):
        try:
            while True:
                msg_type, message = self.result_queue.get_nowait()
                
                self.result_text.config(state=tk.NORMAL)
                
                if msg_type == "file":
                    self.result_text.insert(tk.END, message, "file")
                elif msg_type == "location":
                    self.result_text.insert(tk.END, message, "location")
                elif msg_type == "content":
                    self.result_text.insert(tk.END, message, "content")
                elif msg_type == "error":
                    self.result_text.insert(tk.END, message, "error")
                elif msg_type == "summary":
                    self.result_text.insert(tk.END, message, "summary")
                else:  # info or other
                    self.result_text.insert(tk.END, message)
                
                self.result_text.config(state=tk.DISABLED)
                self.result_text.see(tk.END)
                
                if msg_type == "done":
                    self.progress_bar.stop()
                    self.status_var.set("Search complete")
                    return
                
                self.result_queue.task_done()
        except queue.Empty:
            # No more messages for now, check again in 100ms
            self.root.after(100, self.check_queue)
    
    def search_text_file(self, file_path, search_string):
        """Search for a string in a text file and return matches with line numbers."""
        matches = []
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                for line_num, line in enumerate(file, 1):
                    if search_string.lower() in line.lower():
                        matches.append((f"Line {line_num}", line.strip()))
            return matches
        except Exception as e:
            return [f"Error reading file: {str(e)}"]

    def search_excel_file(self, file_path, search_string):
        """Search for a string in an Excel file and return matches with cell references."""
        matches = []
        try:
            # Try using pandas for faster processing of regular Excel data
            try:
                excel_file = pd.ExcelFile(file_path)
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    
                    # Convert all values to strings for searching
                    for col in df.columns:
                        for row_idx, value in enumerate(df[col]):
                            if value is not None and search_string.lower() in str(value).lower():
                                cell_ref = f"{sheet_name}!{col}{row_idx+2}"  # +2 because of 0-indexing and header
                                matches.append((cell_ref, str(value)))
            
            # If pandas fails, try using openpyxl for more complex Excel files
            except Exception:
                wb = load_workbook(file_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for row in sheet.rows:
                        for cell in row:
                            if cell.value is not None and search_string.lower() in str(cell.value).lower():
                                matches.append((f"{sheet_name}!{cell.coordinate}", str(cell.value)))
                                
            return matches
        except Exception as e:
            return [f"Error reading Excel file: {str(e)}"]

    def search_files_in_folder(self, folder_path, search_string, file_extensions=None):
        """
        Search for a string in all files with specified extensions in a folder.
        Returns a dictionary with file paths as keys and matches as values.
        """
        if file_extensions is None:
            file_extensions = ['.txt', '.xlsx', '.xls', '.csv']
        
        results = {}
        
        for root, _, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                _, extension = os.path.splitext(file)
                
                if extension.lower() in file_extensions:
                    matches = []
                    
                    if extension.lower() in ['.xlsx', '.xls']:
                        matches = self.search_excel_file(file_path, search_string)
                    elif extension.lower() in ['.txt', '.csv']:
                        matches = self.search_text_file(file_path, search_string)
                    
                    if matches:
                        results[file_path] = matches
        
        return results

def main():
    root = tk.Tk()
    app = FileSearchApp(root)
    
    # Configure text tags for formatting
    app.result_text.tag_configure("file", foreground="blue", font=("Arial", 10, "bold"))
    app.result_text.tag_configure("location", foreground="green", font=("Arial", 9))
    app.result_text.tag_configure("content", font=("Arial", 9))
    app.result_text.tag_configure("error", foreground="red", font=("Arial", 9))
    app.result_text.tag_configure("summary", foreground="purple", font=("Arial", 10, "bold"))
    
    root.mainloop()

if __name__ == "__main__":
    main()